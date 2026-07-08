from datetime import date
from io import BytesIO

from openpyxl import Workbook, load_workbook
import pytest
from sqlalchemy import select

from app import database
from app.models import ContractRecord, QuoteRecord
from app.services.excel_importer import import_contract_excel, import_quote_excel
from scripts.create_sample_excels import create_contract_sample, create_quote_sample


QUOTE_HEADERS = [
    "客户名称",
    "国家/地区",
    "型号",
    "替换号码",
    "材质等级",
    "钢板厚度",
    "是否带减震片",
    "包装方式",
    "数量",
    "单价",
    "币种",
    "报价日期",
    "有效期",
    "备注",
]

CONTRACT_HEADERS = [
    "合同编号",
    "客户名称",
    "国家/地区",
    "下单日期",
    "型号",
    "材质等级",
    "包装方式",
    "数量",
    "单价",
    "币种",
    "交货期",
    "付款方式",
    "备注",
]


def build_quote_workbook(rows: list[dict[str, object]]) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "报价单"
    sheet.append(QUOTE_HEADERS)
    for row in rows:
        sheet.append([row.get(header) for header in QUOTE_HEADERS])

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def build_contract_workbook(rows: list[dict[str, object]]) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "合同"
    sheet.append(CONTRACT_HEADERS)
    for row in rows:
        sheet.append([row.get(header) for header in CONTRACT_HEADERS])

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def sample_quote_rows() -> list[dict[str, object]]:
    return [
        {
            "客户名称": "Ahmed",
            "国家/地区": "Libya",
            "型号": "D1234",
            "材质等级": "A+ 半金属",
            "包装方式": "彩盒",
            "数量": 1000,
            "单价": 2.30,
            "币种": "USD",
            "报价日期": date(2026, 7, 1),
        },
        {
            "客户名称": "Ahmed",
            "国家/地区": "Libya",
            "型号": "D5678",
            "材质等级": "AA",
            "包装方式": "彩盒",
            "数量": 500,
            "单价": 3.10,
            "币种": "USD",
            "报价日期": date(2026, 7, 1),
        },
    ]


def sample_contract_rows() -> list[dict[str, object]]:
    return [
        {
            "合同编号": "HT20260707001",
            "客户名称": "Ahmed",
            "国家/地区": "Libya",
            "下单日期": date(2026, 7, 7),
            "型号": "D1234",
            "材质等级": "A+ 半金属",
            "包装方式": "彩盒",
            "数量": 2000,
            "单价": 2.25,
            "币种": "USD",
            "交货期": "30天",
            "付款方式": "T/T",
            "备注": "首单",
        },
        {
            "合同编号": "HT20260707002",
            "客户名称": "Carlos",
            "国家/地区": "Brazil",
            "下单日期": date(2026, 7, 8),
            "型号": "D8888",
            "材质等级": "AAA",
            "包装方式": "中性包装",
            "数量": 800,
            "单价": 4.20,
            "币种": "USD",
        },
    ]


def test_import_quote_excel_with_complete_fields_imports_records(client):
    stream = build_quote_workbook(sample_quote_rows())

    with database.SessionLocal() as db:
        result = import_quote_excel(db, stream, source_file="quotes.xlsx")

    assert result["imported_count"] == 2
    assert result["failed_rows"] == []
    assert result["duplicate_rows"] == []

    with database.SessionLocal() as db:
        records = db.scalars(select(QuoteRecord).order_by(QuoteRecord.part_number)).all()

    assert [record.part_number for record in records] == ["D1234", "D5678"]
    assert records[0].customer_name == "Ahmed"
    assert records[0].country == "Libya"
    assert records[0].material_grade == "A+ 半金属"
    assert records[0].quantity == 1000
    assert str(records[0].unit_price) == "2.3000"
    assert records[0].currency == "USD"
    assert records[0].quote_date == date(2026, 7, 1)
    assert records[0].source_file == "quotes.xlsx"


def test_import_quote_excel_reports_required_customer_name(client):
    row = sample_quote_rows()[0]
    row["客户名称"] = None
    stream = build_quote_workbook([row])

    with database.SessionLocal() as db:
        result = import_quote_excel(db, stream, source_file="missing-customer.xlsx")

    assert result["imported_count"] == 0
    assert result["duplicate_rows"] == []
    assert result["failed_rows"] == [
        {"row_number": 2, "error": "客户名称为必填字段"}
    ]


def test_import_quote_excel_reports_required_part_number(client):
    row = sample_quote_rows()[0]
    row["型号"] = ""
    stream = build_quote_workbook([row])

    with database.SessionLocal() as db:
        result = import_quote_excel(db, stream, source_file="missing-part-number.xlsx")

    assert result["imported_count"] == 0
    assert result["duplicate_rows"] == []
    assert result["failed_rows"] == [{"row_number": 2, "error": "型号为必填字段"}]


def test_import_quote_excel_marks_existing_quote_as_duplicate(client):
    stream = build_quote_workbook([sample_quote_rows()[0]])

    with database.SessionLocal() as db:
        first_result = import_quote_excel(db, stream, source_file="first.xlsx")

    assert first_result["imported_count"] == 1

    duplicate_stream = build_quote_workbook([sample_quote_rows()[0]])
    with database.SessionLocal() as db:
        duplicate_result = import_quote_excel(
            db, duplicate_stream, source_file="duplicate.xlsx"
        )

    assert duplicate_result["imported_count"] == 0
    assert duplicate_result["failed_rows"] == []
    assert duplicate_result["duplicate_rows"] == [
        {
            "row_number": 2,
            "status": "duplicate",
            "message": "客户名称+型号+报价日期已存在",
        }
    ]

    with database.SessionLocal() as db:
        records = db.scalars(select(QuoteRecord)).all()

    assert len(records) == 1
    assert records[0].source_file == "first.xlsx"


def test_import_quote_excel_marks_same_workbook_duplicate_rows(client):
    first_row = sample_quote_rows()[0]
    duplicate_row = sample_quote_rows()[0]
    duplicate_row["备注"] = "duplicate should not overwrite"
    stream = build_quote_workbook([first_row, duplicate_row])

    with database.SessionLocal() as db:
        result = import_quote_excel(db, stream, source_file="same-file.xlsx")

    assert result["imported_count"] == 1
    assert result["failed_rows"] == []
    assert result["duplicate_rows"] == [
        {
            "row_number": 3,
            "status": "duplicate",
            "message": "客户名称+型号+报价日期已存在",
        }
    ]

    with database.SessionLocal() as db:
        records = db.scalars(select(QuoteRecord)).all()

    assert len(records) == 1
    assert records[0].remark is None


def test_import_quote_excel_rejects_fractional_quantity(client):
    row = sample_quote_rows()[0]
    row["数量"] = 1.5
    stream = build_quote_workbook([row])

    with database.SessionLocal() as db:
        result = import_quote_excel(db, stream, source_file="fractional-quantity.xlsx")

    assert result["imported_count"] == 0
    assert result["duplicate_rows"] == []
    assert result["failed_rows"] == [{"row_number": 2, "error": "数量必须为整数"}]

    with database.SessionLocal() as db:
        records = db.scalars(select(QuoteRecord)).all()

    assert records == []


@pytest.mark.parametrize("unit_price", ["NaN", "Infinity", "-Infinity"])
def test_import_quote_excel_rejects_non_finite_unit_price(client, unit_price):
    row = sample_quote_rows()[0]
    row["单价"] = unit_price
    stream = build_quote_workbook([row])

    with database.SessionLocal() as db:
        result = import_quote_excel(db, stream, source_file="nan-price.xlsx")

    assert result["imported_count"] == 0
    assert result["duplicate_rows"] == []
    assert result["failed_rows"] == [{"row_number": 2, "error": "单价必须为有限数字"}]

    with database.SessionLocal() as db:
        records = db.scalars(select(QuoteRecord)).all()

    assert records == []


def test_import_quote_excel_rejects_non_finite_quantity(client):
    row = sample_quote_rows()[0]
    row["数量"] = "Infinity"
    stream = build_quote_workbook([row])

    with database.SessionLocal() as db:
        result = import_quote_excel(db, stream, source_file="infinite-quantity.xlsx")

    assert result["imported_count"] == 0
    assert result["duplicate_rows"] == []
    assert result["failed_rows"] == [{"row_number": 2, "error": "数量必须为有限整数"}]

    with database.SessionLocal() as db:
        records = db.scalars(select(QuoteRecord)).all()

    assert records == []


def test_import_contract_excel_with_complete_fields_imports_records(client):
    stream = build_contract_workbook(sample_contract_rows())

    with database.SessionLocal() as db:
        result = import_contract_excel(db, stream, source_file="contracts.xlsx")

    assert result["imported_count"] == 2
    assert result["failed_rows"] == []
    assert result["duplicate_rows"] == []

    with database.SessionLocal() as db:
        records = db.scalars(
            select(ContractRecord).order_by(ContractRecord.contract_no)
        ).all()

    assert [record.contract_no for record in records] == [
        "HT20260707001",
        "HT20260707002",
    ]
    assert records[0].customer_name == "Ahmed"
    assert records[0].country == "Libya"
    assert records[0].order_date == date(2026, 7, 7)
    assert records[0].part_number == "D1234"
    assert records[0].material_grade == "A+ 半金属"
    assert records[0].packaging == "彩盒"
    assert records[0].quantity == 2000
    assert str(records[0].unit_price) == "2.2500"
    assert records[0].currency == "USD"
    assert records[0].delivery_time == "30天"
    assert records[0].payment_terms == "T/T"
    assert records[0].remark == "首单"
    assert records[0].source_file == "contracts.xlsx"


@pytest.mark.parametrize(
    ("field_name", "expected_error"),
    [
        ("合同编号", "合同编号为必填字段"),
        ("客户名称", "客户名称为必填字段"),
        ("型号", "型号为必填字段"),
    ],
)
def test_import_contract_excel_reports_required_fields(
    client, field_name, expected_error
):
    row = sample_contract_rows()[0]
    row[field_name] = None
    stream = build_contract_workbook([row])

    with database.SessionLocal() as db:
        result = import_contract_excel(db, stream, source_file="missing-contract.xlsx")

    assert result["imported_count"] == 0
    assert result["duplicate_rows"] == []
    assert result["failed_rows"] == [{"row_number": 2, "error": expected_error}]


def test_import_contract_excel_marks_existing_contract_as_duplicate(client):
    stream = build_contract_workbook([sample_contract_rows()[0]])

    with database.SessionLocal() as db:
        first_result = import_contract_excel(db, stream, source_file="first.xlsx")

    assert first_result["imported_count"] == 1

    duplicate_row = sample_contract_rows()[0]
    duplicate_row["型号"] = "D9999"
    duplicate_stream = build_contract_workbook([duplicate_row])
    with database.SessionLocal() as db:
        duplicate_result = import_contract_excel(
            db, duplicate_stream, source_file="duplicate.xlsx"
        )

    assert duplicate_result["imported_count"] == 0
    assert duplicate_result["failed_rows"] == []
    assert duplicate_result["duplicate_rows"] == [
        {
            "row_number": 2,
            "status": "duplicate",
            "message": "合同编号已存在",
        }
    ]

    with database.SessionLocal() as db:
        records = db.scalars(select(ContractRecord)).all()

    assert len(records) == 1
    assert records[0].part_number == "D1234"
    assert records[0].source_file == "first.xlsx"


def test_import_contract_excel_marks_same_workbook_duplicate_rows(client):
    first_row = sample_contract_rows()[0]
    duplicate_row = sample_contract_rows()[0]
    duplicate_row["备注"] = "duplicate should not overwrite"
    stream = build_contract_workbook([first_row, duplicate_row])

    with database.SessionLocal() as db:
        result = import_contract_excel(db, stream, source_file="same-file-contracts.xlsx")

    assert result["imported_count"] == 1
    assert result["failed_rows"] == []
    assert result["duplicate_rows"] == [
        {
            "row_number": 3,
            "status": "duplicate",
            "message": "合同编号已存在",
        }
    ]

    with database.SessionLocal() as db:
        records = db.scalars(select(ContractRecord)).all()

    assert len(records) == 1
    assert records[0].remark == "首单"


def test_import_contracts_endpoint_accepts_excel_file(client):
    stream = build_contract_workbook(sample_contract_rows())

    response = client.post(
        "/api/imports/contracts",
        files={
            "file": (
                "contracts.xlsx",
                stream.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    assert response.json() == {
        "imported_count": 2,
        "failed_rows": [],
        "duplicate_rows": [],
    }


def test_import_quotes_endpoint_accepts_excel_file(client):
    stream = build_quote_workbook(sample_quote_rows())

    response = client.post(
        "/api/imports/quotes",
        files={
            "file": (
                "quotes.xlsx",
                stream.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    assert response.json() == {
        "imported_count": 2,
        "failed_rows": [],
        "duplicate_rows": [],
    }


def test_create_quote_sample_generates_required_rows(tmp_path):
    sample_path = create_quote_sample(tmp_path / "sample_quotes.xlsx")

    workbook = load_workbook(sample_path, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    assert sample_path.exists()
    assert rows[0] == tuple(QUOTE_HEADERS)
    assert rows[1][0:3] == ("Ahmed", "Libya", "D1234")
    assert rows[1][4] == "A+ 半金属"
    assert rows[1][8:11] == (1000, 2.30, "USD")
    assert rows[2][0:3] == ("Ahmed", "Libya", "D5678")
    assert rows[2][4] == "AA"
    assert rows[2][8:11] == (500, 3.10, "USD")


def test_create_contract_sample_generates_required_rows(tmp_path):
    sample_path = create_contract_sample(tmp_path / "sample_contracts.xlsx")

    workbook = load_workbook(sample_path, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    assert sample_path.exists()
    assert rows[0] == tuple(CONTRACT_HEADERS)
    assert rows[1][0:3] == ("HT20260707001", "Ahmed", "Libya")
    assert rows[1][3].date() == date(2026, 7, 7)
    assert rows[1][4] == "D1234"
    assert rows[1][5:10] == ("A+ 半金属", "彩盒", 2000, 2.25, "USD")
    assert rows[2][0:3] == ("HT20260707002", "Carlos", "Brazil")
    assert rows[2][3].date() == date(2026, 7, 8)
    assert rows[2][4] == "D8888"
    assert rows[2][5:10] == ("AAA", "中性包装", 800, 4.20, "USD")


def test_import_quotes_endpoint_rejects_invalid_excel_file(client):
    response = client.post(
        "/api/imports/quotes",
        files={"file": ("broken.xlsx", b"not an excel workbook", "text/plain")},
    )

    assert response.status_code == 400
    assert response.json() == {"detail": "无法读取 Excel 文件，请上传有效的 .xlsx 文件"}


def test_import_contracts_endpoint_rejects_invalid_excel_file(client):
    response = client.post(
        "/api/imports/contracts",
        files={"file": ("broken.xlsx", b"not an excel workbook", "text/plain")},
    )

    assert response.status_code == 400
    assert response.json() == {"detail": "无法读取 Excel 文件，请上传有效的 .xlsx 文件"}
