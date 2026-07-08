from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import BinaryIO
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import ContractRecord, QuoteRecord


QUOTE_FIELD_MAPPING = {
    "客户名称": "customer_name",
    "国家/地区": "country",
    "型号": "part_number",
    "替换号码": "replacement_numbers",
    "材质等级": "material_grade",
    "钢板厚度": "steel_thickness",
    "是否带减震片": "has_shim",
    "包装方式": "packaging",
    "数量": "quantity",
    "单价": "unit_price",
    "币种": "currency",
    "报价日期": "quote_date",
    "有效期": "valid_until",
    "备注": "remark",
}

QUOTE_REQUIRED_FIELDS = {
    "customer_name": "客户名称",
    "country": "国家/地区",
    "part_number": "型号",
    "material_grade": "材质等级",
    "packaging": "包装方式",
    "quantity": "数量",
    "unit_price": "单价",
    "currency": "币种",
    "quote_date": "报价日期",
}

CONTRACT_FIELD_MAPPING = {
    "合同编号": "contract_no",
    "客户名称": "customer_name",
    "国家/地区": "country",
    "下单日期": "order_date",
    "型号": "part_number",
    "材质等级": "material_grade",
    "包装方式": "packaging",
    "数量": "quantity",
    "单价": "unit_price",
    "币种": "currency",
    "交货期": "delivery_time",
    "付款方式": "payment_terms",
    "备注": "remark",
}

CONTRACT_REQUIRED_FIELDS = {
    "contract_no": "合同编号",
    "customer_name": "客户名称",
    "country": "国家/地区",
    "order_date": "下单日期",
    "part_number": "型号",
    "material_grade": "材质等级",
    "packaging": "包装方式",
    "quantity": "数量",
    "unit_price": "单价",
    "currency": "币种",
}

TRUE_VALUES = {"是", "有", "yes", "y", "true", "1", "带"}
FALSE_VALUES = {"否", "无", "no", "n", "false", "0", "不带"}


class InvalidExcelFileError(ValueError):
    pass


def import_quote_excel(
    db: Session, file: BinaryIO, source_file: str | None = None
) -> dict[str, object]:
    try:
        workbook = load_workbook(file, data_only=True)
    except (BadZipFile, InvalidFileException, OSError, KeyError, ValueError) as exc:
        raise InvalidExcelFileError("无法读取 Excel 文件，请上传有效的 .xlsx 文件") from exc
    sheet = workbook.active
    headers = [_normalize_header(cell.value) for cell in sheet[1]]
    header_indexes = {
        QUOTE_FIELD_MAPPING[header]: index
        for index, header in enumerate(headers)
        if header in QUOTE_FIELD_MAPPING
    }

    imported_count = 0
    failed_rows: list[dict[str, object]] = []
    duplicate_rows: list[dict[str, object]] = []
    imported_quote_keys: set[tuple[str, str, date]] = set()

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if _is_empty_row(row):
            continue

        values = {
            field_name: row[index] if index < len(row) else None
            for field_name, index in header_indexes.items()
        }

        error = _validate_required_fields(values, QUOTE_REQUIRED_FIELDS)
        if error is not None:
            failed_rows.append({"row_number": row_number, "error": error})
            continue

        try:
            quote_data = _convert_quote_values(values)
        except ValueError as exc:
            failed_rows.append({"row_number": row_number, "error": str(exc)})
            continue

        quote_key = (
            quote_data["customer_name"],
            quote_data["part_number"],
            quote_data["quote_date"],
        )
        if quote_key in imported_quote_keys or _quote_exists(
            db,
            customer_name=quote_data["customer_name"],
            part_number=quote_data["part_number"],
            quote_date=quote_data["quote_date"],
        ):
            duplicate_rows.append(
                {
                    "row_number": row_number,
                    "status": "duplicate",
                    "message": "客户名称+型号+报价日期已存在",
                }
            )
            continue

        db.add(QuoteRecord(**quote_data, source_file=source_file))
        imported_quote_keys.add(quote_key)
        imported_count += 1

    if imported_count:
        db.commit()

    return {
        "imported_count": imported_count,
        "failed_rows": failed_rows,
        "duplicate_rows": duplicate_rows,
    }


def import_contract_excel(
    db: Session, file: BinaryIO, source_file: str | None = None
) -> dict[str, object]:
    try:
        workbook = load_workbook(file, data_only=True)
    except (BadZipFile, InvalidFileException, OSError, KeyError, ValueError) as exc:
        raise InvalidExcelFileError("无法读取 Excel 文件，请上传有效的 .xlsx 文件") from exc
    sheet = workbook.active
    headers = [_normalize_header(cell.value) for cell in sheet[1]]
    header_indexes = {
        CONTRACT_FIELD_MAPPING[header]: index
        for index, header in enumerate(headers)
        if header in CONTRACT_FIELD_MAPPING
    }

    imported_count = 0
    failed_rows: list[dict[str, object]] = []
    duplicate_rows: list[dict[str, object]] = []
    imported_contract_nos: set[str] = set()

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if _is_empty_row(row):
            continue

        values = {
            field_name: row[index] if index < len(row) else None
            for field_name, index in header_indexes.items()
        }

        error = _validate_required_fields(values, CONTRACT_REQUIRED_FIELDS)
        if error is not None:
            failed_rows.append({"row_number": row_number, "error": error})
            continue

        try:
            contract_data = _convert_contract_values(values)
        except ValueError as exc:
            failed_rows.append({"row_number": row_number, "error": str(exc)})
            continue

        contract_no = contract_data["contract_no"]
        if contract_no in imported_contract_nos or _contract_exists(db, contract_no):
            duplicate_rows.append(
                {
                    "row_number": row_number,
                    "status": "duplicate",
                    "message": "合同编号已存在",
                }
            )
            continue

        db.add(ContractRecord(**contract_data, source_file=source_file))
        imported_contract_nos.add(contract_no)
        imported_count += 1

    if imported_count:
        db.commit()

    return {
        "imported_count": imported_count,
        "failed_rows": failed_rows,
        "duplicate_rows": duplicate_rows,
    }


def _normalize_header(value: object) -> str:
    return str(value).strip() if value is not None else ""


def _is_empty_row(row: tuple[object, ...]) -> bool:
    return all(value is None or str(value).strip() == "" for value in row)


def _validate_required_fields(
    values: dict[str, object], required_fields: dict[str, str]
) -> str | None:
    for field_name, label in required_fields.items():
        value = values.get(field_name)
        if value is None or str(value).strip() == "":
            return f"{label}为必填字段"
    return None


def _convert_quote_values(values: dict[str, object]) -> dict[str, object]:
    return {
        "customer_name": _to_text(values.get("customer_name")),
        "country": _to_text(values.get("country")),
        "part_number": _to_text(values.get("part_number")),
        "replacement_numbers": _to_text(values.get("replacement_numbers")),
        "material_grade": _to_text(values.get("material_grade")),
        "steel_thickness": _to_text(values.get("steel_thickness")),
        "has_shim": _to_bool(values.get("has_shim")),
        "packaging": _to_text(values.get("packaging")),
        "quantity": _to_int(values.get("quantity"), "数量"),
        "unit_price": _to_decimal(values.get("unit_price"), "单价"),
        "currency": _to_text(values.get("currency")),
        "quote_date": _to_date(values.get("quote_date"), "报价日期"),
        "valid_until": _to_date(values.get("valid_until"), "有效期"),
        "remark": _to_text(values.get("remark")),
    }


def _convert_contract_values(values: dict[str, object]) -> dict[str, object]:
    return {
        "contract_no": _to_text(values.get("contract_no")),
        "customer_name": _to_text(values.get("customer_name")),
        "country": _to_text(values.get("country")),
        "order_date": _to_date(values.get("order_date"), "下单日期"),
        "part_number": _to_text(values.get("part_number")),
        "material_grade": _to_text(values.get("material_grade")),
        "packaging": _to_text(values.get("packaging")),
        "quantity": _to_int(values.get("quantity"), "数量"),
        "unit_price": _to_decimal(values.get("unit_price"), "单价"),
        "currency": _to_text(values.get("currency")),
        "delivery_time": _to_text(values.get("delivery_time")),
        "payment_terms": _to_text(values.get("payment_terms")),
        "remark": _to_text(values.get("remark")),
    }


def _to_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _to_bool(value: object) -> bool | None:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, bool):
        return value
    normalized = str(value).strip().lower()
    if normalized in TRUE_VALUES:
        return True
    if normalized in FALSE_VALUES:
        return False
    return None


def _to_int(value: object, label: str) -> int:
    if isinstance(value, bool):
        raise ValueError(f"{label}必须为整数")
    try:
        decimal_value = Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        raise ValueError(f"{label}必须为整数") from None
    if not decimal_value.is_finite():
        raise ValueError(f"{label}必须为有限整数")
    if decimal_value != decimal_value.to_integral_value():
        raise ValueError(f"{label}必须为整数")
    return int(decimal_value)


def _to_decimal(value: object, label: str) -> Decimal:
    try:
        decimal_value = Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        raise ValueError(f"{label}必须为数字") from None
    if not decimal_value.is_finite():
        raise ValueError(f"{label}必须为有限数字")
    return decimal_value


def _to_date(value: object, label: str) -> date | None:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for date_format in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, date_format).date()
        except ValueError:
            continue
    raise ValueError(f"{label}必须为日期")


def _quote_exists(
    db: Session, customer_name: str, part_number: str, quote_date: date
) -> bool:
    existing_id = db.scalar(
        select(QuoteRecord.id).where(
            QuoteRecord.customer_name == customer_name,
            QuoteRecord.part_number == part_number,
            QuoteRecord.quote_date == quote_date,
        )
    )
    return existing_id is not None


def _contract_exists(db: Session, contract_no: str) -> bool:
    existing_id = db.scalar(
        select(ContractRecord.id).where(ContractRecord.contract_no == contract_no)
    )
    return existing_id is not None
